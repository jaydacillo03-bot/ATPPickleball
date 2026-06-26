from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

wb=Workbook()
ws=wb.active
ws.title="Tournament Scoresheet"

thin=Side(style="thin")
med=Side(style="medium")

def box(r1,c1,r2,c2):
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):
            cell=ws.cell(r,c)
            cell.border=Border(left=thin,right=thin,top=thin,bottom=thin)

ws.merge_cells("A1:L1")
t=ws["A1"]; t.value="PICKLEBALL TOURNAMENT SCORESHEET"
t.font=Font(size=18,bold=True)
t.alignment=Alignment(horizontal="center")

info=[
("A3","Tournament"),("F3","Date"),
("A4","Division"),("F4","Court"),
("A5","Round"),("F5","Match #"),
("A6","Referee"),("F6","Format (11/15/21)")
]
for cell,label in info:
    ws[cell]=label
    ws[cell].font=Font(bold=True)

# teams
ws.merge_cells("A8:F8"); ws["A8"]="TEAM A"; ws["A8"].font=Font(bold=True)
ws.merge_cells("G8:L8"); ws["G8"]="TEAM B"; ws["G8"].font=Font(bold=True)
for rng in [("A9","F11"),("G9","L11")]:
    start,end=rng
    # names
ws["A9"]="Player 1"; ws["A10"]="Player 2"
ws["G9"]="Player 1"; ws["G10"]="Player 2"
box(8,1,11,6); box(8,7,11,12)

# scoring table
headers=["Rally","Serving","Server\n(1/2)","Side","A","B","Timeout","End?","Notes"]
start_row=13
for i,h in enumerate(headers,1):
    c=ws.cell(start_row,i)
    c.value=h
    c.font=Font(bold=True)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.fill=PatternFill("solid",fgColor="D9EAD3")
    c.border=Border(left=med,right=med,top=med,bottom=med)

for r in range(start_row+1,start_row+41):
    ws.cell(r,1).value=r-start_row
    for c in range(1,len(headers)+1):
        ws.cell(r,c).border=Border(left=thin,right=thin,top=thin,bottom=thin)
        ws.cell(r,c).alignment=Alignment(horizontal="center")

summary_row=56
fields=[("A56","Final Score"),("A57","Winner"),("D56","Team A"),("F56","Team B"),
("A59","Timeouts A"),("D59","Timeouts B"),("A61","Remarks")]
for cell,val in fields:
    ws[cell]=val
    ws[cell].font=Font(bold=True)

for col in range(1,13):
    ws.column_dimensions[get_column_letter(col)].width=14

path="/Users/joceldacillo/Project/Professional_Pickleball_Tournament_Scoresheet.xlsx"
wb.save(path)

print(path)
STDOUT/STDERR
/mnt/data/Professional_Pickleball_Tournament_Scoresheet.xlsx